import openpyxl
import math
from funkcijas import *
import numpy as np

rind_colon = sezon_input("Dati csv failā ir rindā (rakstiet '1') vai kolonā (rakstiet '2')","Lūdzu ievadiet '1' vai '2'",'1','2','exfgchjpect3')
if rind_colon == "1":
    with open("data.csv", "r" , encoding="utf8") as f:
        next(f)
        next(f)
        row1 = next(f)
        exdate  = row1.rstrip().split(";")
        row2 = next(f)
        data  = row2.rstrip().split(";")
    index = 0
    for elem in data:
        try:
            value = float(elem)
            break
        except ValueError:
            index = index + 1
    del exdate[0:index]
    del data[0:index]
    date = []
    for elem in exdate:
        a = elem.replace('"','')
        date.append(a)

if rind_colon == "2":
    data = []
    exdate = []
    date = []
    with open("data.csv","r", encoding="utf8") as f:
        next(f)
        next(f)
        next(f)
        for line in f:
            row = line.rstrip().split(";")
            data.append(float(row[-1]))
            exdate.append(str(row[-2]))
    for elem in exdate:
        a = elem.replace('"','')
        date.append(a)


wb = openpyxl.Workbook()
ws=wb.active

ws['A1'] = "Slidosas vidējas metode"
slid_data = VidSlid(ws,date,data,3,3)
ws[Xlsx_Letter(6+len(data)//2) + "1"] = "Vidēja kvadratiska klūda"
slid_data_klud = KvadKlud(ws,date,data, 3,3,3+len(data)//2,3+len(data))


ws['A' + str(len(data) + 8)] = "Eksponenciālas noguldīsānas metode"
ekspon_data = EksponNoglud(ws,date,data,3,len(data) + 10)
ws[Xlsx_Letter(6+9) + str(len(data) + 8)] = "Vidēja kvadratiska klūda"
ekspon_data_klud = KvadKlud(ws,date,data, 3,len(data) + 10,3+9,len(data) + 10+len(data))


ws['A' + str(2*len(data) + 15)] = "Tendences metode"
trend_data = TendencesMetode(ws,date,data,3,2*len(data) + 17)
ws[Xlsx_Letter(6+11) + str(2*len(data) + 15)] = "Vidēja kvadratiska klūda"
trend_data_klud = KvadKlud(ws,date,data, 4,2*len(data) + 17,3+11,2*len(data) + 17+len(data))


Sezon_papild = sezon_input("Kāds laika periods ir starp datiem? Ja mēnesis: ievadi '1', ja kvartile: ievadi '2', ja gads: ievadi '3'","Lūgums ievadīt '1', '2' vai '3'","1","2","3")
if Sezon_papild == "1":
    data_sadal = 12
if Sezon_papild == "2":
    data_sadal = 4
if Sezon_papild == "3":
    print("Sezonāla metode navar būt izpildīta ar dotajiem datiem")

if Sezon_papild == "1" or Sezon_papild == "2":
    first_sezon = first_sezon_input("Kāds kārta numurs ir 1. periodam?","Lūdzu ievadiet skaitli no 1 līdz ",data_sadal)
    ws['A' + str(3*len(data) + 22)] = "Sezonālais modelis"
    sezon_data = SezonMet(ws,date,data,data_sadal,int(first_sezon),3,3*len(data) + 24)
    ws[Xlsx_Letter(6+5) + str(3*len(data) + 22)] = "Vidēja kvadratiska klūda"
    sezon_data_klud = KvadKlud(ws,date,data, 7,3*len(data) + 24,8,3*len(data) + 24+len(data))

slid_min_index = np.array(slid_data_klud[1]).argmin()
expon_min_index = np.array(ekspon_data_klud[1]).argmin()
trnd_min_index = np.array(trend_data_klud[1]).argmin()

min_klud = []
best_prognoz = []

if Sezon_papild == "1" or Sezon_papild == "2":
    rezult_papild = 4*len(data) + 29
else:
    rezult_papild = 3*len(data) + 22


ws['A' + str(rezult_papild)] = "Rezultāti"
ws['A' + str(rezult_papild + 1)] = "Metode"
ws['B' + str(rezult_papild + 1)] = "Modelis"
ws['C' + str(rezult_papild + 1)] = "Kļūda"
ws['D' + str(rezult_papild + 1)] = "Prognoze"

ws['A' + str(rezult_papild + 2)] = "Slidosas vidējas metode"
ws['B' + str(rezult_papild + 2)] = str(ws[Xlsx_Letter(3+int(slid_min_index)) + "2"].value)
ws['C' + str(rezult_papild + 2)] = slid_data_klud[1][slid_min_index]
min_klud.append(slid_data_klud[1][slid_min_index])
ws['D' + str(rezult_papild + 2)] = slid_data[1][slid_min_index]
best_prognoz.append(slid_data[1][slid_min_index])

ws['A' + str(rezult_papild + 3)] = "Eksponenciālas noguldīsānas metode"
ws['B' + str(rezult_papild + 3)] = str(ws[Xlsx_Letter(3+int(expon_min_index)) + str(len(data) + 9)].value)
ws['C' + str(rezult_papild + 3)] = ekspon_data_klud[1][expon_min_index]
min_klud.append(ekspon_data_klud[1][expon_min_index])
ws['D' + str(rezult_papild + 3)] = ekspon_data[1][expon_min_index]
best_prognoz.append(ekspon_data[1][expon_min_index])

ws['A' + str(rezult_papild + 4)] = "Tendences metode"
ws['B' + str(rezult_papild + 4)] = str(ws[Xlsx_Letter(4+int(trnd_min_index)) + str(2*len(data) + 16)].value)
ws['C' + str(rezult_papild + 4)] = trend_data_klud[1][trnd_min_index]
min_klud.append(trend_data_klud[1][trnd_min_index])
ws['D' + str(rezult_papild + 4)] = trend_data[1][trnd_min_index]
best_prognoz.append(trend_data[1][trnd_min_index])

if Sezon_papild == "1" or Sezon_papild == "2":
    ws['A' + str(rezult_papild + 5)] = "Sezonālais modelis"
    ws['B' + str(rezult_papild + 5)] = "Sezonālais modelis"
    ws['C' + str(rezult_papild + 5)] = sezon_data_klud[1][0]
    min_klud.append(sezon_data_klud[1][0])
    ws['D' + str(rezult_papild + 5)] = sezon_data[1]
    best_prognoz.append(sezon_data[1])

min_kluda_index = np.array(min_klud).argmin()
if min_kluda_index == 0:
    best_metode = "Slidosas vidējas metode"
    best_modelis = str(ws[Xlsx_Letter(3+int(slid_min_index)) + "2"].value)
    best_next = best_prognoz[0]

if min_kluda_index == 1:
    best_metode = "Eksponenciālas noguldīsānas metode"
    best_modelis = str(ws[Xlsx_Letter(3+int(expon_min_index)) + str(len(data) + 9)].value)
    best_next = best_prognoz[1]

if min_kluda_index == 2:
    best_metode = "Tendences metode"
    best_modelis = str(ws[Xlsx_Letter(4+int(trnd_min_index)) + str(2*len(data) + 16)].value)
    best_next = best_prognoz[2]

if min_kluda_index == 3:
    best_metode = "Sezonālais modelis"
    best_modelis = "Sezonālais modelis"
    best_next = best_prognoz[3]


if_save = sezon_input("Vai vēlaties saglabāt xlsx failu? Ja jā: ievadiet 'yes', ja nē: ievadiet 'no'","Lūdzu ievadiet 'yes' vai 'no'","yes","no","dftg687y7uhj")
if if_save == "yes":
    try_to_name(wb,"Ievadiet faila nosaukumu (bez .xlsx)","Lūdzu nosaukamā neizmantojiet '/', '\', '?', '%', '*', ':', '|', '”', '<', '>' un citus simbolus, kurus nevar izmantot kā faila nosaukumu")
wb.close()

print("Rezultātu ar vismazāko kvadratīstko kļūdu (" + str(min_klud[min_kluda_index]) + ") paradīja "+ best_metode + " ar " + best_modelis + ".")
print("Tās ir " + str(best_next) + ".")
print("No tā seko, ka nākāma perioda vērtība būs no " + str(best_next-pow(min_klud[min_kluda_index],0.5)) + " līdz " + str(best_next+pow(min_klud[min_kluda_index],0.5)) + ".")